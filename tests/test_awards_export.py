from __future__ import annotations

import io
import unittest
from unittest import mock

import pandas as pd
from openpyxl import load_workbook

from src.analysis import contractors_combined_detail, filter_transactions_for_contractors
from src.awards_export import (
    EXPORT_COLUMNS,
    awards_export_xlsx,
    build_awards_export_frame,
    build_leaderboard_export_frame,
    leaderboard_export_xlsx,
)
from src.state import FilterSnapshot
from tests.test_analysis import sample_transactions


class ContractorSelectionTests(unittest.TestCase):
    def test_filter_transactions_for_multiple_contractors(self):
        scoped = filter_transactions_for_contractors(
            sample_transactions(),
            ["Acme Global LLC", "Beta Services Inc."],
        )
        self.assertEqual(int(scoped["canonical_contractor"].nunique()), 2)
        self.assertEqual(float(scoped["federal_action_obligation"].sum()), 1400.0)

    def test_combined_detail_aggregates_selected_contractors(self):
        detail = contractors_combined_detail(
            sample_transactions(),
            ["Acme Global LLC", "Beta Services Inc."],
        )
        self.assertEqual(detail["unique_awards"], 2)
        self.assertEqual(detail["obligations"], 1400.0)


class AwardsExportTests(unittest.TestCase):
    def test_export_frame_omits_redundant_url_columns(self):
        from src.analysis import award_table, filter_transactions

        awards = award_table(filter_transactions(sample_transactions(), FilterSnapshot(agency="Department of State")))
        export_df = build_awards_export_frame(awards)
        self.assertEqual(list(export_df.columns), EXPORT_COLUMNS)
        self.assertNotIn("Recipient Profile URL", export_df.columns)
        self.assertNotIn("Award URL", export_df.columns)

    def test_excel_hyperlinks_live_on_contractor_and_award_id_only(self):
        from src.analysis import award_table, filter_transactions

        awards = award_table(filter_transactions(sample_transactions(), FilterSnapshot(agency="Department of State")))
        with mock.patch(
            "src.awards_export.usaspending_recipient_profile_url",
            return_value="https://www.usaspending.gov/recipient/test/latest",
        ):
            xlsx_bytes = awards_export_xlsx(awards)
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        self.assertNotIn("Recipient Profile URL", headers)
        self.assertNotIn("Award URL", headers)
        contractor_col = headers.index("Contractor") + 1
        award_id_col = headers.index("Award ID") + 1
        contractor_cell = worksheet.cell(row=2, column=contractor_col)
        award_cell = worksheet.cell(row=2, column=award_id_col)
        self.assertEqual(contractor_cell.hyperlink.target, "https://www.usaspending.gov/recipient/test/latest")
        self.assertTrue(str(award_cell.hyperlink.target).startswith("https://www.usaspending.gov/award/"))


class LeaderboardExportTests(unittest.TestCase):
    def test_leaderboard_export_frame_formats_dates(self):
        from src.analysis import recent_wins_leaderboard

        leaderboard = recent_wins_leaderboard(sample_transactions())
        export_df = build_leaderboard_export_frame(leaderboard)
        self.assertIn("Contractor Name", export_df.columns)
        self.assertIn("Primary UEI", export_df.columns)
        if not export_df.empty and "Most Recent Win" in export_df.columns:
            recent = export_df.iloc[0]["Most Recent Win"]
            self.assertIsNotNone(recent)

    def test_leaderboard_excel_hyperlinks_contractor_name(self):
        from src.analysis import recent_wins_leaderboard

        leaderboard = recent_wins_leaderboard(sample_transactions())
        with mock.patch(
            "src.awards_export.usaspending_recipient_profile_url",
            return_value="https://www.usaspending.gov/recipient/test/latest",
        ):
            xlsx_bytes = leaderboard_export_xlsx(leaderboard, worksheet_title="Recent Service Wins")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        worksheet = workbook.active
        self.assertEqual(worksheet.title, "Recent Service Wins")
        headers = [cell.value for cell in worksheet[1]]
        contractor_col = headers.index("Contractor Name") + 1
        if worksheet.max_row >= 2:
            contractor_cell = worksheet.cell(row=2, column=contractor_col)
            self.assertEqual(contractor_cell.hyperlink.target, "https://www.usaspending.gov/recipient/test/latest")


if __name__ == "__main__":
    unittest.main()

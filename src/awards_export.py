from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

from .utils import usaspending_recipient_profile_url


EXPORT_COLUMNS = [
    "Contractor",
    "Recipient UEI",
    "Award ID",
    "Description",
    "Obligations in Scope",
    "Current Award Value",
    "Award Ceiling",
    "Performance Location",
    "Awarding Office",
    "Funding Office",
]


def build_awards_export_frame(awards: pd.DataFrame) -> pd.DataFrame:
    if awards is None or awards.empty:
        return pd.DataFrame(columns=EXPORT_COLUMNS)
    rows = []
    for row in awards.to_dict("records"):
        rows.append(
            {
                "Contractor": str(row.get("Contractor") or ""),
                "Recipient UEI": str(row.get("Recipient UEI") or ""),
                "Award ID": row.get("Award ID"),
                "Description": row.get("Description"),
                "Obligations in Scope": row.get("Obligations in Scope"),
                "Current Award Value": row.get("Current Award Value"),
                "Award Ceiling": row.get("Award Ceiling"),
                "Performance Location": row.get("Performance Location"),
                "Awarding Office": row.get("Awarding Office"),
                "Funding Office": row.get("Funding Office"),
            }
        )
    return pd.DataFrame(rows)


def awards_export_xlsx(awards: pd.DataFrame) -> bytes:
    export_df = build_awards_export_frame(awards)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Top Relevant Awards"
    headers = list(export_df.columns)
    worksheet.append(headers)
    link_font = Font(color="0563C1", underline="single")
    contractor_col = headers.index("Contractor") + 1
    award_id_col = headers.index("Award ID") + 1
    for export_row, source_row in zip(export_df.to_dict("records"), awards.to_dict("records")):
        worksheet.append([export_row.get(header) for header in headers])
        row_idx = worksheet.max_row
        profile_url = usaspending_recipient_profile_url(
            str(source_row.get("Recipient UEI") or ""),
            str(source_row.get("Contractor") or ""),
        )
        award_url = str(source_row.get("USAspending Award Link") or "")
        contractor_cell = worksheet.cell(row=row_idx, column=contractor_col)
        if profile_url:
            contractor_cell.hyperlink = profile_url
            contractor_cell.font = link_font
        award_cell = worksheet.cell(row=row_idx, column=award_id_col)
        if award_url:
            award_cell.hyperlink = award_url
            award_cell.font = link_font
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def build_leaderboard_export_frame(leaderboard: pd.DataFrame) -> pd.DataFrame:
    if leaderboard is None or leaderboard.empty:
        return pd.DataFrame()
    export = leaderboard.copy()
    for date_column in ("Most Recent Win", "Most Recent Action Date"):
        if date_column in export.columns:
            export[date_column] = pd.to_datetime(export[date_column], errors="coerce").dt.date
    return export


def leaderboard_export_xlsx(leaderboard: pd.DataFrame, *, worksheet_title: str) -> bytes:
    export_df = build_leaderboard_export_frame(leaderboard)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = worksheet_title[:31]
    headers = list(export_df.columns)
    worksheet.append(headers)
    if export_df.empty:
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    link_font = Font(color="0563C1", underline="single")
    contractor_col = headers.index("Contractor Name") + 1 if "Contractor Name" in headers else None
    for export_row, source_row in zip(export_df.to_dict("records"), leaderboard.to_dict("records")):
        worksheet.append([export_row.get(header) for header in headers])
        if contractor_col is None:
            continue
        row_idx = worksheet.max_row
        profile_url = usaspending_recipient_profile_url(
            str(source_row.get("Primary UEI") or ""),
            str(source_row.get("Contractor Name") or ""),
        )
        contractor_cell = worksheet.cell(row=row_idx, column=contractor_col)
        if profile_url:
            contractor_cell.hyperlink = profile_url
            contractor_cell.font = link_font
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
